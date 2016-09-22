from openpyxl import load_workbook


def colToExcel(col):
    excelCol = str()
    div = col
    while div:
        (div, mod) = divmod(div - 1, 26)
        excelCol = chr(mod + 65) + excelCol
    return excelCol



def draw_column(excel_file):
    wb=load_workbook(filename=excel_file,read_only=True)
    ws = wb.active
    column_count = ws.max_column
    column_name=[]

    print "column_count=",column_count

    for column in range(1, column_count + 1):
        column_name.append(colToExcel(column))

    res=[]# first row name of the columns

    for i in column_name:
        cell_name = "{}{}".format(i, 1)
        res.append(ws[cell_name].value)


    return res,column_name
