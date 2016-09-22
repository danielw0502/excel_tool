from openpyxl import load_workbook
from os.path import exists
from openpyxl import Workbook



def write_column(dest_file,original_file,column_name):

    if not exists(dest_file):
        kb=Workbook()
        kb.save(dest_file)

    kb=load_workbook(filename=dest_file)
    ks=kb.active
    column_write_count = ks.max_column

    #firstly write the original file name to the dest_file
    ks.cell(row=1, column=column_write_count + 1).value = original_file

    #secondly write the data of the column you select to the dest file
    wb = load_workbook(filename=original_file, read_only=True)
    ws = wb.active
    row_count = ws.max_row

    column_len=len(column_name)


    #for jj in range(0, column_len):



    for row in range(1, row_count + 1):
        com = column_write_count

        for column in column_name:

            cell_name = "{}{}".format(column, row)


            ks.cell(row=row, column=com + 2).value = ws[cell_name].value

            com += 1

    kb.save(filename=dest_file)