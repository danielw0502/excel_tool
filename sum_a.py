#calculate the eNB rate
#2016-10-1
#Author:wxd

#from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.utils import coordinate_from_string,column_index_from_string
#import pdb

wb=load_workbook(filename='a.xlsx')

ws=wb.active

#pdb.set_trace()

#cell_range=ws['A1':'B15']

flag=0
sum_a=0
res=0

count=0
a=0

for i in range(1,16):
    for j in range(1,3):
        if count==0:
            a=ws.cell(row=i,column=j).value
        elif count==1 and ws.cell(row=i,column=j).value!=0:
            #pdb.set_trace()
            sum_a+=a
        elif count==1 and ws.cell(row=i,column=j).value==0:
            res+=sum_a
            #xy=coordinate_from_string(cell)
            #col=column_index_from_string(xy[0])
            #row=xy[1]
            ws.cell(row=i,column=j+1).value=sum_a
            sum_a=0
        count+=1
    count=0
    a=0
wb.save(filename='a.xlsx')
print res

